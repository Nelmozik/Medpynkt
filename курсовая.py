from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, make_response
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from functools import wraps
import os
import io
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.config['SECRET_KEY'] = 'medpunkt-secret-key-2024'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)


# ========== МОДЕЛИ ==========
class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password = db.Column(db.String(50), nullable=False)
    full_name = db.Column(db.String(200), nullable=False)
    position = db.Column(db.String(200), default='Медсестра')
    role = db.Column(db.String(20), default='nurse')
    phone = db.Column(db.String(20), default='')
    email = db.Column(db.String(100), default='')
    avatar = db.Column(db.String(10), default='👤')


class Student(db.Model):
    __tablename__ = 'students'
    id = db.Column(db.Integer, primary_key=True)
    full_name = db.Column(db.String(200), nullable=False)
    group_name = db.Column(db.String(50))
    birth_date = db.Column(db.Date)
    phone = db.Column(db.String(20))
    medical_notes = db.Column(db.Text, default='')
    allergies = db.Column(db.Text, default='')
    blood_type = db.Column(db.String(5), default='')
    vaccinations = db.relationship('VaccinationLog', backref='student', lazy=True, cascade="all, delete-orphan")


class Vaccine(db.Model):
    __tablename__ = 'vaccines'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text)
    period_years = db.Column(db.Integer, default=0)
    period_months = db.Column(db.Integer, default=0)
    mandatory = db.Column(db.Boolean, default=True)


class VaccinationLog(db.Model):
    __tablename__ = 'vaccination_log'
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('students.id'), nullable=False)
    vaccine_id = db.Column(db.Integer, db.ForeignKey('vaccines.id'), nullable=False)
    date_given = db.Column(db.Date, nullable=False)
    notes = db.Column(db.Text)
    vaccine = db.relationship('Vaccine', backref='logs')


class Settings(db.Model):
    __tablename__ = 'settings'
    id = db.Column(db.Integer, primary_key=True)
    college_name = db.Column(db.String(200), default='Колледж')
    reminder_days = db.Column(db.Integer, default=30)
    theme = db.Column(db.String(20), default='medical')
    auto_backup = db.Column(db.Boolean, default=False)
    language = db.Column(db.String(10), default='ru')
    notifications_enabled = db.Column(db.Boolean, default=True)


# ========== ДЕКОРАТОРЫ ==========
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('🔒 Пожалуйста, войдите в систему', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)

    return decorated_function


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('🔒 Пожалуйста, войдите в систему', 'warning')
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            flash('⛔ Доступ запрещён. Только для администратора.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)

    return decorated_function


# ========== ФУНКЦИИ ==========
def get_settings():
    settings = Settings.query.first()
    if not settings:
        settings = Settings()
        db.session.add(settings)
        db.session.commit()
    return settings


def get_current_user():
    if 'user_id' in session:
        return User.query.get(session['user_id'])
    return None


def get_theme():
    # Сначала проверяем куки, потом настройки из БД
    theme = request.cookies.get('theme')
    if not theme:
        settings = get_settings()
        theme = settings.theme
    return theme if theme in ['medical', 'dark', 'light'] else 'medical'


def calculate_due_date(last_date, period_years, period_months):
    if last_date is None:
        return None
    return last_date + relativedelta(years=period_years, months=period_months)


def get_status_color(due_date):
    if due_date is None:
        return ""
    today = datetime.now().date()
    settings = get_settings()
    if due_date < today:
        return "table-danger"
    elif due_date <= today + timedelta(days=settings.reminder_days):
        return "table-warning"
    else:
        return "table-success"


# ========== АВТОРИЗАЦИЯ ==========
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        user = User.query.filter_by(username=username, password=password).first()

        if user:
            session['user_id'] = user.id
            session['role'] = user.role
            session['user_name'] = user.full_name
            flash(f'👋 Добро пожаловать, {user.full_name}!', 'success')
            return redirect(url_for('index'))
        else:
            flash('❌ Неверный логин или пароль', 'danger')

    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    flash('👋 Вы вышли из системы', 'info')
    return redirect(url_for('login'))


# ========== СМЕНА ТЕМЫ ==========
@app.route('/set-theme/<theme>')
def set_theme(theme):
    if theme not in ['medical', 'dark', 'light']:
        theme = 'medical'

    # Сохраняем в БД
    settings = get_settings()
    settings.theme = theme
    db.session.commit()

    # Сохраняем в куки
    resp = make_response(redirect(request.referrer or url_for('index')))
    resp.set_cookie('theme', theme, max_age=365 * 24 * 60 * 60)  # на год
    return resp


# ========== РЕЗЕРВНОЕ КОПИРОВАНИЕ ==========
@app.route('/backup')
@admin_required
def backup_database():
    import shutil
    from datetime import datetime

    # Папка для бэкапов
    backup_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'backups')
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir)

    # Путь к текущей базе
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'database.db')

    # Проверяем, существует ли база
    if not os.path.exists(db_path):
        flash('❌ Файл базы данных не найден!', 'danger')
        return redirect(url_for('app_settings'))

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_path = os.path.join(backup_dir, f'backup_{timestamp}.db')

    try:
        shutil.copy2(db_path, backup_path)
        flash(f'💾 Резервная копия создана: backup_{timestamp}.db', 'success')
    except Exception as e:
        flash(f'❌ Ошибка при создании копии: {str(e)}', 'danger')

    return redirect(url_for('app_settings'))

# ========== СТАТИСТИКА ==========
@app.route('/statistics')
@login_required
def statistics():
    settings = get_settings()
    current_user = get_current_user()

    # Общая статистика
    total_students = Student.query.count()
    total_vaccines = Vaccine.query.count()
    total_vaccinations = VaccinationLog.query.count()

    # Статистика по группам
    from sqlalchemy import func
    group_stats = db.session.query(
        Student.group_name,
        func.count(Student.id)
    ).filter(Student.group_name != '').group_by(Student.group_name).all()

    # Статистика по вакцинам
    vaccine_stats = []
    for vaccine in Vaccine.query.all():
        count = VaccinationLog.query.filter_by(vaccine_id=vaccine.id).count()
        vaccine_stats.append({
            'name': vaccine.name,
            'count': count,
            'mandatory': vaccine.mandatory
        })

    # Последние прививки
    recent_vaccinations = VaccinationLog.query.order_by(
        VaccinationLog.date_given.desc()
    ).limit(10).all()

    return render_template('statistics.html',
                           settings=settings,
                           current_user=current_user,
                           total_students=total_students,
                           total_vaccines=total_vaccines,
                           total_vaccinations=total_vaccinations,
                           group_stats=group_stats,
                           vaccine_stats=vaccine_stats,
                           recent_vaccinations=recent_vaccinations)


# ========== ПОИСК ==========
@app.route('/search')
@login_required
def search():
    query = request.args.get('q', '').strip()
    settings = get_settings()
    current_user = get_current_user()

    results = []
    if query:
        results = Student.query.filter(
            Student.full_name.ilike(f'%{query}%')
        ).order_by(Student.full_name).all()

    return render_template('search.html',
                           settings=settings,
                           current_user=current_user,
                           query=query,
                           results=results)


# ========== ОСНОВНЫЕ МАРШРУТЫ ==========
@app.route('/')
@login_required
def index():
    settings = get_settings()
    current_user = get_current_user()
    theme = get_theme()
    students = Student.query.all()
    vaccines = Vaccine.query.all()

    total_students = len(students)
    alerts_count = 0
    overdue_count = 0  # НОВОЕ: счётчик просроченных
    vaccinated_count = 0

    today = datetime.now().date()
    alert_date = today + timedelta(days=settings.reminder_days)

    for student in students:
        has_alert = False
        has_overdue = False  # НОВОЕ: флаг просрочки
        for vaccine in vaccines:
            last_vax = VaccinationLog.query.filter_by(
                student_id=student.id,
                vaccine_id=vaccine.id
            ).order_by(VaccinationLog.date_given.desc()).first()

            if last_vax:
                due_date = calculate_due_date(last_vax.date_given, vaccine.period_years, vaccine.period_months)
                if due_date and due_date < today:  # НОВОЕ: проверка на просрочку
                    has_overdue = True
                if due_date and due_date <= alert_date:
                    has_alert = True
            else:
                has_alert = True
                has_overdue = True  # НОВОЕ: нет данных = тоже просрочено

        if has_overdue:
            overdue_count += 1
        if has_alert:
            alerts_count += 1
        else:
            vaccinated_count += 1

    return render_template('index.html',
                           settings=settings,
                           current_user=current_user,
                           theme=theme,
                           total_students=total_students,
                           alerts_count=alerts_count,
                           overdue_count=overdue_count,  # НОВОЕ
                           vaccinated_count=vaccinated_count,
                           now=datetime.now())


@app.route('/students')
@login_required
def students():
    settings = get_settings()
    current_user = get_current_user()
    theme = get_theme()
    students = Student.query.order_by(Student.full_name).all()
    vaccines = Vaccine.query.all()

    student_summary = []
    for student in students:
        summary = {'student': student, 'vaccines_status': []}
        for vaccine in vaccines:
            last_vax = VaccinationLog.query.filter_by(
                student_id=student.id,
                vaccine_id=vaccine.id
            ).order_by(VaccinationLog.date_given.desc()).first()

            if last_vax:
                last_date = last_vax.date_given
                due_date = calculate_due_date(last_date, vaccine.period_years, vaccine.period_months)
            else:
                last_date = None
                due_date = None

            summary['vaccines_status'].append({
                'vaccine': vaccine,
                'last_date': last_date,
                'due_date': due_date,
                'status_class': get_status_color(due_date)
            })
        student_summary.append(summary)

    return render_template('students.html', settings=settings, current_user=current_user, theme=theme,
                           student_summary=student_summary, vaccines=vaccines)


@app.route('/report')
@login_required
def report():
    settings = get_settings()
    current_user = get_current_user()
    theme = get_theme()
    today = datetime.now().date()
    alert_date = today + timedelta(days=settings.reminder_days)

    students = Student.query.all()
    vaccines = Vaccine.query.all()

    alerts = []
    for student in students:
        for vaccine in vaccines:
            last_vax = VaccinationLog.query.filter_by(
                student_id=student.id,
                vaccine_id=vaccine.id
            ).order_by(VaccinationLog.date_given.desc()).first()

            if last_vax:
                last_date = last_vax.date_given
                due_date = calculate_due_date(last_date, vaccine.period_years, vaccine.period_months)
            else:
                due_date = student.birth_date + relativedelta(years=vaccine.period_years,
                                                              months=vaccine.period_months) if student.birth_date else None

            if due_date and due_date <= alert_date:
                days_left = (due_date - today).days
                if days_left < 0:
                    status_text = f"ПРОСРОЧЕНО на {abs(days_left)} дн."
                    alert_class = "alert-danger"
                elif days_left == 0:
                    status_text = "СЕГОДНЯ"
                    alert_class = "alert-danger"
                else:
                    status_text = f"Через {days_left} дн."
                    alert_class = "alert-warning"

                alerts.append({
                    'student': student,
                    'vaccine': vaccine,
                    'due_date': due_date,
                    'status_text': status_text,
                    'alert_class': alert_class
                })

    alerts.sort(key=lambda x: (x['alert_class'] == 'alert-danger', x['alert_class'] == 'alert-warning'))

    return render_template('report.html', settings=settings, current_user=current_user, theme=theme, alerts=alerts,
                           today=today)


@app.route('/student/add', methods=['GET', 'POST'])
@login_required
def add_student():
    settings = get_settings()
    current_user = get_current_user()
    theme = get_theme()
    if request.method == 'POST':
        student = Student(
            full_name=request.form['full_name'],
            group_name=request.form.get('group_name', ''),
            birth_date=datetime.strptime(request.form['birth_date'], '%Y-%m-%d').date() if request.form.get(
                'birth_date') else None,
            phone=request.form.get('phone', ''),
            medical_notes=request.form.get('medical_notes', ''),
            allergies=request.form.get('allergies', ''),
            blood_type=request.form.get('blood_type', '')
        )
        db.session.add(student)
        db.session.commit()
        flash('✅ Студент успешно добавлен!', 'success')
        return redirect(url_for('students'))
    return render_template('add_student.html', settings=settings, current_user=current_user, theme=theme)


@app.route('/student/edit/<int:student_id>', methods=['GET', 'POST'])
@login_required
def edit_student(student_id):
    settings = get_settings()
    current_user = get_current_user()
    theme = get_theme()
    student = Student.query.get_or_404(student_id)

    if request.method == 'POST':
        student.full_name = request.form['full_name']
        student.group_name = request.form.get('group_name', '')
        student.birth_date = datetime.strptime(request.form['birth_date'], '%Y-%m-%d').date() if request.form.get(
            'birth_date') else None
        student.phone = request.form.get('phone', '')
        student.medical_notes = request.form.get('medical_notes', '')
        student.allergies = request.form.get('allergies', '')
        student.blood_type = request.form.get('blood_type', '')
        db.session.commit()
        flash('✅ Данные студента обновлены!', 'success')
        return redirect(url_for('students'))

    return render_template('edit_student.html', settings=settings, current_user=current_user, theme=theme,
                           student=student)


@app.route('/student/delete/<int:student_id>')
@admin_required
def delete_student(student_id):
    student = Student.query.get_or_404(student_id)
    db.session.delete(student)
    db.session.commit()
    flash('🗑️ Студент удалён!', 'success')
    return redirect(url_for('students'))


@app.route('/vaccination/add/<int:student_id>', methods=['GET', 'POST'])
@login_required
def add_vaccination(student_id):
    settings = get_settings()
    current_user = get_current_user()
    theme = get_theme()
    student = Student.query.get_or_404(student_id)
    vaccines = Vaccine.query.all()

    if request.method == 'POST':
        vax = VaccinationLog(
            student_id=student_id,
            vaccine_id=int(request.form['vaccine_id']),
            date_given=datetime.strptime(request.form['date_given'], '%Y-%m-%d').date() if request.form.get(
                'date_given') else datetime.now().date(),
            notes=request.form.get('notes', '')
        )
        db.session.add(vax)
        db.session.commit()
        flash('💉 Прививка зарегистрирована!', 'success')
        return redirect(url_for('students'))

    return render_template('add_vaccination.html', settings=settings, current_user=current_user, theme=theme,
                           student=student, vaccines=vaccines)


@app.route('/vaccines')
@login_required
def vaccines_list():
    settings = get_settings()
    current_user = get_current_user()
    theme = get_theme()
    vaccines = Vaccine.query.all()
    return render_template('vaccines.html', settings=settings, current_user=current_user, theme=theme,
                           vaccines=vaccines)


@app.route('/vaccine/add', methods=['GET', 'POST'])
@admin_required
def add_vaccine():
    settings = get_settings()
    current_user = get_current_user()
    theme = get_theme()
    if request.method == 'POST':
        vaccine = Vaccine(
            name=request.form['name'],
            description=request.form.get('description', ''),
            period_years=int(request.form.get('period_years', 0)),
            period_months=int(request.form.get('period_months', 0)),
            mandatory=bool(request.form.get('mandatory', True))
        )
        db.session.add(vaccine)
        db.session.commit()
        flash('💉 Вакцина добавлена!', 'success')
        return redirect(url_for('vaccines_list'))
    return render_template('add_vaccine.html', settings=settings, current_user=current_user, theme=theme)


@app.route('/vaccine/edit/<int:vaccine_id>', methods=['GET', 'POST'])
@admin_required
def edit_vaccine(vaccine_id):
    settings = get_settings()
    current_user = get_current_user()
    theme = get_theme()
    vaccine = Vaccine.query.get_or_404(vaccine_id)

    if request.method == 'POST':
        vaccine.name = request.form['name']
        vaccine.description = request.form.get('description', '')
        vaccine.period_years = int(request.form.get('period_years', 0))
        vaccine.period_months = int(request.form.get('period_months', 0))
        vaccine.mandatory = bool(request.form.get('mandatory', True))
        db.session.commit()
        flash('✅ Данные вакцины обновлены!', 'success')
        return redirect(url_for('vaccines_list'))

    return render_template('edit_vaccine.html', settings=settings, current_user=current_user, theme=theme,
                           vaccine=vaccine)


@app.route('/vaccine/delete/<int:vaccine_id>')
@admin_required
def delete_vaccine(vaccine_id):
    vaccine = Vaccine.query.get_or_404(vaccine_id)
    db.session.delete(vaccine)
    db.session.commit()
    flash('🗑️ Вакцина удалена!', 'success')
    return redirect(url_for('vaccines_list'))


@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    settings = get_settings()
    current_user = get_current_user()
    theme = get_theme()

    if request.method == 'POST':
        current_user.full_name = request.form.get('full_name', current_user.full_name)
        current_user.phone = request.form.get('phone', '')
        current_user.email = request.form.get('email', '')
        current_user.avatar = request.form.get('avatar', current_user.avatar)
        if request.form.get('new_password'):
            current_user.password = request.form['new_password']
        db.session.commit()
        session['user_name'] = current_user.full_name
        flash('👤 Профиль обновлён!', 'success')
        return redirect(url_for('profile'))

    return render_template('profile.html', settings=settings, current_user=current_user, theme=theme)


@app.route('/users')
@admin_required
def users_list():
    settings = get_settings()
    current_user = get_current_user()
    theme = get_theme()
    users = User.query.all()
    return render_template('users.html', settings=settings, current_user=current_user, theme=theme, users=users)


@app.route('/user/add', methods=['GET', 'POST'])
@admin_required
def add_user():
    settings = get_settings()
    current_user = get_current_user()
    theme = get_theme()

    if request.method == 'POST':
        user = User(
            username=request.form['username'],
            password=request.form['password'],
            full_name=request.form['full_name'],
            position=request.form.get('position', 'Медсестра'),
            role=request.form.get('role', 'nurse'),
            phone=request.form.get('phone', ''),
            email=request.form.get('email', ''),
            avatar=request.form.get('avatar', '👤')
        )
        db.session.add(user)
        db.session.commit()
        flash('✅ Пользователь создан!', 'success')
        return redirect(url_for('users_list'))

    return render_template('add_user.html', settings=settings, current_user=current_user, theme=theme)


@app.route('/user/delete/<int:user_id>')
@admin_required
def delete_user(user_id):
    if user_id == session.get('user_id'):
        flash('⚠️ Нельзя удалить самого себя!', 'danger')
        return redirect(url_for('users_list'))
    user = User.query.get_or_404(user_id)
    db.session.delete(user)
    db.session.commit()
    flash('🗑️ Пользователь удалён!', 'success')
    return redirect(url_for('users_list'))


@app.route('/settings', methods=['GET', 'POST'])
@admin_required
def app_settings():
    settings = get_settings()
    current_user = get_current_user()
    theme = get_theme()

    if request.method == 'POST':
        settings.college_name = request.form.get('college_name', 'Колледж')
        settings.reminder_days = int(request.form.get('reminder_days', 30))
        settings.theme = request.form.get('theme', 'medical')
        settings.auto_backup = bool(request.form.get('auto_backup', False))
        settings.notifications_enabled = bool(request.form.get('notifications_enabled', True))
        db.session.commit()

        # Обновляем куки темы
        resp = make_response(redirect(url_for('app_settings')))
        resp.set_cookie('theme', settings.theme, max_age=365 * 24 * 60 * 60)

        flash('⚙️ Настройки сохранены!', 'success')
        return resp

    return render_template('settings.html', settings=settings, current_user=current_user, theme=theme)


# Экспорт в Excel
@app.route('/export/excel')
@login_required
def export_excel():
    settings = get_settings()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Прививки"

    headers = ['ФИО', 'Группа', 'Телефон', 'Группа крови', 'Аллергии']
    vaccines = Vaccine.query.all()
    for v in vaccines:
        headers.append(v.name)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    students = Student.query.all()
    for row, student in enumerate(students, 2):
        ws.cell(row=row, column=1, value=student.full_name)
        ws.cell(row=row, column=2, value=student.group_name)
        ws.cell(row=row, column=3, value=student.phone)
        ws.cell(row=row, column=4, value=student.blood_type)
        ws.cell(row=row, column=5, value=student.allergies)

        col = 6
        for vaccine in vaccines:
            last_vax = VaccinationLog.query.filter_by(
                student_id=student.id,
                vaccine_id=vaccine.id
            ).order_by(VaccinationLog.date_given.desc()).first()

            if last_vax:
                next_date = calculate_due_date(last_vax.date_given, vaccine.period_years, vaccine.period_months)
                today = datetime.now().date()
                if next_date and next_date > today + timedelta(days=settings.reminder_days):
                    status = "✅ "
                elif next_date and next_date >= today:
                    status = "⚠️ "
                else:
                    status = "❌ "
                value = f"{status}{last_vax.date_given.strftime('%d.%m.%Y')}"
            else:
                value = "❌ Нет данных"

            cell = ws.cell(row=row, column=col, value=value)
            if "❌" in value:
                cell.fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
            elif "⚠️" in value:
                cell.fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")

            col += 1

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Отчёт_прививки_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
    )


@app.route('/export/report/excel')
@login_required
def export_report_excel():
    settings = get_settings()
    today = datetime.now().date()
    alert_date = today + timedelta(days=settings.reminder_days)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчёт на сегодня"

    headers = ['ФИО студента', 'Группа', 'Телефон', 'Вакцина', 'Статус', 'Дата']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    row = 2
    students = Student.query.all()
    vaccines = Vaccine.query.all()

    for student in students:
        for vaccine in vaccines:
            last_vax = VaccinationLog.query.filter_by(
                student_id=student.id,
                vaccine_id=vaccine.id
            ).order_by(VaccinationLog.date_given.desc()).first()

            if last_vax:
                last_date = last_vax.date_given
                due_date = calculate_due_date(last_date, vaccine.period_years, vaccine.period_months)
            else:
                due_date = student.birth_date + relativedelta(years=vaccine.period_years,
                                                              months=vaccine.period_months) if student.birth_date else None

            if due_date and due_date <= alert_date:
                days_left = (due_date - today).days
                status_text = f"ПРОСРОЧЕНО на {abs(days_left)} дн." if days_left < 0 else f"Через {days_left} дн."

                ws.cell(row=row, column=1, value=student.full_name)
                ws.cell(row=row, column=2, value=student.group_name)
                ws.cell(row=row, column=3, value=student.phone)
                ws.cell(row=row, column=4, value=vaccine.name)
                ws.cell(row=row, column=5, value=status_text)
                ws.cell(row=row, column=6, value=due_date.strftime('%d.%m.%Y'))

                if days_left < 0:
                    for c in range(1, 7):
                        ws.cell(row=row, column=c).fill = PatternFill(start_color="f8d7da", end_color="f8d7da",
                                                                      fill_type="solid")

                row += 1

    for column in ws.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Отчёт_на_сегодня_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
    )


# ========== ИНИЦИАЛИЗАЦИЯ ==========
def init_db():
    with app.app_context():
        db.create_all()

        if User.query.count() == 0:
            admin = User(
                username='admin',
                password='admin123',
                full_name='Иванова Мария Петровна',
                position='Администратор',
                role='admin',
                phone='+7 (999) 123-45-67',
                email='admin@college.ru',
                avatar='👑'
            )
            nurse = User(
                username='nurse',
                password='nurse123',
                full_name='Петрова Анна Сергеевна',
                position='Медицинская сестра',
                role='nurse',
                phone='+7 (999) 765-43-21',
                email='nurse@college.ru',
                avatar='👩‍⚕️'
            )
            db.session.add(admin)
            db.session.add(nurse)
            db.session.commit()

        if Vaccine.query.count() == 0:
            vaccines = [
                Vaccine(name='АДС-М (дифтерия, столбняк)', description='Ревакцинация взрослых каждые 10 лет',
                        period_years=10, mandatory=True),
                Vaccine(name='Гепатит B', description='Вакцинация взрослых до 55 лет', mandatory=True),
                Vaccine(name='Грипп', description='Ежегодная вакцинация', period_years=1, mandatory=False),
                Vaccine(name='Корь', description='Вакцинация взрослых до 35 лет', mandatory=True),
            ]
            for v in vaccines:
                db.session.add(v)
            db.session.commit()

        if Settings.query.count() == 0:
            settings = Settings(
                college_name='Мой Колледж',
                reminder_days=30,
                theme='medical',
                auto_backup=False,
                notifications_enabled=True
            )
            db.session.add(settings)
            db.session.commit()

        if Student.query.count() == 0:
            test_student = Student(
                full_name='Иванов Иван Иванович',
                group_name='ИС-302',
                birth_date=datetime(2005, 5, 15).date(),
                phone='+79001234567',
                blood_type='II+',
                allergies='Нет'
            )
            db.session.add(test_student)
            db.session.commit()


if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=5000)